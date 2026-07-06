import pandas as pd
import numpy as np
import json
import io
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

from app.services.analysis_settings import DEFAULT_ANALYSIS_SETTINGS, normalize_analysis_settings

class StudyAnalysisService:
    def __init__(self):
        self.rng = np.random.default_rng(123)
        
        # Constants
        self.PANEL_COL = "Panelist"
        self.RATING_COL = "Rating"
        self.RESPONSE_TIME_COL = "ResponseTime"
        self.GENDER_COL = "Gender"
        self.AGE_COL = "Age"
        self.TASK_COL = "Task"
        
        self.AGE_BINS = [
            "13-17", "18-24", "25-34", "35-44", 
            "45-54", "55-64", "65+"
        ]
        
        # Styles
        self.bold_font = Font(bold=True)
        self.header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        self.blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        self.MAX_WIDTH = 45
        self.letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    def _resolve_analysis_options(self, analysis_options: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        return normalize_analysis_settings(analysis_options or DEFAULT_ANALYSIS_SETTINGS)

    def _include_intercept(self, analysis_options: Dict[str, Any]) -> bool:
        return bool((analysis_options.get("regression") or {}).get("include_intercept", True))

    def _build_y_vector(
        self,
        ratings: np.ndarray,
        mode: str,
        analysis_options: Dict[str, Any],
        jitter_seed: Optional[int] = None,
    ) -> np.ndarray:
        if mode == "RESPONSE":
            return np.clip(ratings.astype(float), None, 7.0)

        scoring = analysis_options["top"] if mode == "TOP" else analysis_options["bottom"]
        hundred = set(scoring.get("hundred") or [])
        Y = np.array(
            [100.0 if int(r) in hundred else 0.0 for r in ratings],
            dtype=float,
        )
        rng = self.rng if jitter_seed is None else np.random.default_rng(jitter_seed)
        Y = Y + rng.uniform(-0.5, 0.5, size=Y.shape) * 1e-5
        return Y

    def generate_report(
        self,
        df: pd.DataFrame,
        study_data: Dict[str, Any],
        analysis_options: Optional[Dict[str, Any]] = None,
        filters: Optional[Dict[str, Any]] = None,
    ) -> io.BytesIO:
        """
        Generates the Excel report from the DataFrame and Study Data.
        Returns a BytesIO object containing the Excel file.

        When ``filters`` is provided and non-empty, panelists are subset first via
        ``_filter_df_by_filters`` and all sheets are computed on that cohort only.
        """
        analysis_options = self._resolve_analysis_options(analysis_options)
        normalized_filters = self._normalize_filters_dict(filters)
        filters_active = self._filters_are_active(normalized_filters)

        if filters_active:
            df = self._filter_df_by_filters(
                df,
                age_groups=normalized_filters.get("age_groups"),
                genders=normalized_filters.get("genders"),
                classification_filters=normalized_filters.get("classification_filters"),
            )
            df = df.sort_values([self.PANEL_COL, self.TASK_COL]).reset_index(drop=True) if not df.empty else df
            if df.empty:
                raise ValueError("No respondents match the applied filters.")

        # 1. Preprocess Data
        # Handle NA values if not already handled
        # df is passed in, assuming it's already reasonably clean from response service
        # but we might need to handle specific NA representations if any
        
        # Extract metadata
        elements_json = study_data.get("elements", [])
        categories_json = study_data.get("categories", [])
        
        # 2. Element Metadata & Column Mapping
        element_meta = []
        for el in elements_json:
            cat_obj = el.get("category", {})
            # Fallback if category object is not fully populated but we have category_id
            if not cat_obj and el.get("category_id"):
                 # Find category in categories_json
                 found_cat = next((c for c in categories_json if c.get("id") == el.get("category_id")), {})
                 cat_obj = found_cat
            
            cat_name = cat_obj.get("name")
            el_name = el.get("name")
            
            if not cat_name or not el_name:
                continue
                
            # Construct expected column name (handling the mismatch)
            # analysis_v2 expected: f"{cat_name}_{el_name}"
            # response.py generates: f"{cat_name}-{el_name}".replace('_', '-').replace(' ', '-')
            
            # We need to find which column in df corresponds to this element
            # Let's try to match flexible
            
            # Expected "clean" name for internal logic
            internal_col_name = f"{cat_name}_{el_name}"
            
            # Try to find matching column in DF
            # 1. Try exact match (unlikely given the mismatch)
            # 2. Try hyphenated version
            # 3. Try hyphenated + safe replacements
            
            candidates = [
                internal_col_name,
                f"{cat_name}-{el_name}",
                f"{cat_name}-{el_name}".replace('_', '-').replace(' ', '-'),
                f"{cat_name}_{el_name}".replace(' ', '_')
            ]
            
            actual_col = None
            for cand in candidates:
                if cand in df.columns:
                    actual_col = cand
                    break
            
            if actual_col:
                element_meta.append({
                    "csv_col": actual_col, # Use the actual column name in DF
                    "category_name": cat_name,
                    "element_name": el_name,
                    "category_order": cat_obj.get("order", 0),
                })

        # Keep only columns that exist and deduplicate (some elements may have same name)
        element_cols_raw = [m["csv_col"] for m in element_meta]
        # Deduplicate while preserving order
        seen = set()
        element_cols = []
        for col in element_cols_raw:
            if col not in seen:
                element_cols.append(col)
                seen.add(col)
        
        # Maps
        col_to_catname = {m["csv_col"]: m["category_name"] for m in element_meta}
        col_to_eltname = {m["csv_col"]: m["element_name"] for m in element_meta}
        
        # Category ordering
        cat_order = {}
        for m in element_meta:
            name = m["category_name"]
            order = m["category_order"]
            if name not in cat_order or order < cat_order[name]:
                cat_order[name] = order
        sorted_categories = sorted(cat_order.keys(), key=lambda c: cat_order[c])

        # 3. Classification Columns
        # In response.py, classification columns are named by Question Text
        # We can identify them by excluding known columns
        known_cols = {self.PANEL_COL, self.RATING_COL, self.RESPONSE_TIME_COL, 
                      self.GENDER_COL, self.AGE_COL, self.TASK_COL}
        known_cols.update(element_cols)
        
        # Also exclude "session_id" etc if present
        classification_cols = []
        # Heuristic: Columns between Age and Task? Or just use study_data
        # Using study_data is safer
        class_qs = study_data.get("classification_questions", [])
        for q in class_qs:
            q_text = q.get("question_text")
            if q_text and q_text in df.columns:
                classification_cols.append(q_text)

        selected_classification_filters = normalized_filters.get("classification_filters") or {}
        if filters_active and selected_classification_filters:
            selected_questions = set(selected_classification_filters.keys())
            unselected_classification_cols = [
                col for col in classification_cols if col not in selected_questions
            ]
            if unselected_classification_cols:
                df = df.drop(columns=unselected_classification_cols, errors="ignore")
            classification_cols = [
                col for col in classification_cols if col in selected_questions
            ]
                
        # 4. Run Analysis
        # 4a. Panel-level Regressions
        coef_table_T = self._run_panel_regressions(df, element_cols, "TOP", analysis_options)
        coef_table_B = self._run_panel_regressions(df, element_cols, "BOTTOM", analysis_options)
        coef_table_R = self._run_panel_regressions(df, element_cols, "RESPONSE", analysis_options)
        
        # 4b. Aggregations (Means & Groups)
        # Base Size
        base_size = df[self.PANEL_COL].nunique()
        
        # Means
        element_means_T = coef_table_T[element_cols].mean(axis=0).round().astype(int)
        element_means_B = coef_table_B[element_cols].mean(axis=0).round().astype(int)
        element_means_R = coef_table_R[element_cols].mean(axis=0) # Float for RT
        
        # Groups
        # Need to pass df to build groups because we need Gender/Age/Class info mapped to Panelist
        gender_groups_T = self._build_gender_groups(coef_table_T, df, element_cols)
        gender_groups_B = self._build_gender_groups(coef_table_B, df, element_cols)
        gender_groups_R = self._build_gender_groups(coef_table_R, df, element_cols)
        
        age_groups_T = self._build_age_groups(coef_table_T, df, element_cols)
        age_groups_B = self._build_age_groups(coef_table_B, df, element_cols)
        age_groups_R = self._build_age_groups(coef_table_R, df, element_cols)
        
        class_groups_T = self._build_class_groups(coef_table_T, df, element_cols, classification_cols)
        class_groups_B = self._build_class_groups(coef_table_B, df, element_cols, classification_cols)
        class_groups_R = self._build_class_groups(coef_table_R, df, element_cols, classification_cols)
        
        # 4c. Pooled Regressions (Intercepts)
        intercepts_T = self._run_pooled_regression(df, element_cols, "TOP", analysis_options)
        intercepts_B = self._run_pooled_regression(df, element_cols, "BOTTOM", analysis_options)
        intercepts_R = self._run_pooled_regression(df, element_cols, "RESPONSE", analysis_options)
        
        coef_threshold_T = intercepts_T.get("threshold")
        coef_threshold_B = intercepts_B.get("threshold")
        coef_threshold_R = intercepts_R.get("threshold")
        
        # 5. Excel Generation
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        # 5a. Front Page & Info
        self._create_front_page(wb, study_data)
        self._create_info_block(wb, study_data)
        self._create_raw_data_sheet(wb, df)
        
        # 5b. Overall Sheets
        self._create_overall_sheet(wb, "(T) Overall", element_cols, sorted_categories, col_to_catname, col_to_eltname, 
                                   element_means_T, base_size, coef_threshold_T, self.green_fill, round_vals=True)
        self._create_overall_sheet(wb, "(B) Overall", element_cols, sorted_categories, col_to_catname, col_to_eltname, 
                                   element_means_B, base_size, coef_threshold_B, self.red_fill, round_vals=True)
        self._create_overall_sheet(wb, "(R) Overall", element_cols, sorted_categories, col_to_catname, col_to_eltname, 
                                   element_means_R, base_size, coef_threshold_R, self.blue_fill, round_vals=False)
                                   
        # 5c. Mindsets (Clustering)
        # Run clustering on T coefficients
        X_T = coef_table_T[element_cols].to_numpy(dtype=float)
        n_samples = X_T.shape[0]
        
        # Handle clustering based on available samples
        # Need at least k samples to create k clusters
        if n_samples >= 2:
            labels_2_T = self._custom_kmeans_pearson(X_T, k=2, seed=101)
        else:
            labels_2_T = np.zeros(n_samples, dtype=int)
            
        if n_samples >= 3:
            labels_3_T = self._custom_kmeans_pearson(X_T, k=3, seed=202)
        else:
            labels_3_T = np.zeros(n_samples, dtype=int)

        self._create_mindset_sheet(wb, "(T) Mindsets", coef_table_T, element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   base_size, coef_threshold_T, self.green_fill, labels_2_T, labels_3_T, round_vals=True)
        self._create_mindset_sheet(wb, "(B) Mindsets", coef_table_B, element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   base_size, coef_threshold_B, self.red_fill, labels_2_T, labels_3_T, round_vals=True)
        self._create_mindset_sheet(wb, "(R) Mindsets", coef_table_R, element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   base_size, coef_threshold_R, self.blue_fill, labels_2_T, labels_3_T, round_vals=False)

        # 5d. Gender Sheets
        self._create_segment_sheet(wb, "(T) Gender", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   gender_groups_T, coef_threshold_T, self.green_fill, round_vals=True)
        self._create_segment_sheet(wb, "(B) Gender", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   gender_groups_B, coef_threshold_B, self.red_fill, round_vals=True)
        self._create_segment_sheet(wb, "(R) Gender", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   gender_groups_R, coef_threshold_R, self.blue_fill, round_vals=False)

        # 5e. Age Sheets
        self._create_segment_sheet(wb, "(T) Age", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   age_groups_T, coef_threshold_T, self.green_fill, round_vals=True, segment_order=self.AGE_BINS)
        self._create_segment_sheet(wb, "(B) Age", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   age_groups_B, coef_threshold_B, self.red_fill, round_vals=True, segment_order=self.AGE_BINS)
        self._create_segment_sheet(wb, "(R) Age", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                   age_groups_R, coef_threshold_R, self.blue_fill, round_vals=False, segment_order=self.AGE_BINS)

        # 5f. Classification Sheets
        self._create_classification_sheet(wb, "(T) Classification Questions", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                          class_groups_T, coef_threshold_T, self.green_fill, round_vals=True)
        self._create_classification_sheet(wb, "(B) Classification Questions", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                          class_groups_B, coef_threshold_B, self.red_fill, round_vals=True)
        self._create_classification_sheet(wb, "(R) Classification Questions", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                          class_groups_R, coef_threshold_R, self.blue_fill, round_vals=False)

        # 5g. Combined Sheets
        self._create_combined_sheet(wb, "(T) Combined", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                    base_size, element_means_T, gender_groups_T, age_groups_T, class_groups_T,
                                    coef_threshold_T, self.green_fill, round_vals=True)
        self._create_combined_sheet(wb, "(B) Combined", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                    base_size, element_means_B, gender_groups_B, age_groups_B, class_groups_B,
                                    coef_threshold_B, self.red_fill, round_vals=True)
        self._create_combined_sheet(wb, "(R) Combined", element_cols, sorted_categories, col_to_catname, col_to_eltname,
                                    base_size, element_means_R, gender_groups_R, age_groups_R, class_groups_R,
                                    coef_threshold_R, self.blue_fill, round_vals=False)

        # 5h. Intercepts Sheets
        self._create_intercepts_sheet(
            wb, "(T) Intercepts", intercepts_T["df"], coef_threshold_T, self.green_fill,
            intercept=intercepts_T.get("intercept"), t_intercept=intercepts_T.get("t_intercept"),
        )
        self._create_intercepts_sheet(
            wb, "(B) Intercepts", intercepts_B["df"], coef_threshold_B, self.red_fill,
            intercept=intercepts_B.get("intercept"), t_intercept=intercepts_B.get("t_intercept"),
        )
        self._create_intercepts_sheet(
            wb, "(R) Intercepts", intercepts_R["df"], coef_threshold_R, self.blue_fill,
            intercept=intercepts_R.get("intercept"), t_intercept=intercepts_R.get("t_intercept"),
        )

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_json_report(
        self,
        df: pd.DataFrame,
        study_data: Dict[str, Any],
        include_raw_data: bool = True,
        analysis_options: Optional[Dict[str, Any]] = None,
        filters: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Generates the JSON report from the DataFrame and Study Data.
        Returns a dictionary with sheet names as keys and their data as values.

        When ``filters`` is provided and non-empty, panelists are subset first via
        ``_filter_df_by_filters`` and all sections (dashboard_summary, Overall,
        Age, Gender, Mindsets, Classification, etc.) are computed on that cohort only.
        """
        analysis_options = self._resolve_analysis_options(analysis_options)
        normalized_filters = self._normalize_filters_dict(filters)
        filters_active = self._filters_are_active(normalized_filters)

        rows_before = int(len(df))
        panelists_before = int(df[self.PANEL_COL].nunique()) if self.PANEL_COL in df.columns and not df.empty else 0

        if filters_active:
            df = self._filter_df_by_filters(
                df,
                age_groups=normalized_filters.get("age_groups"),
                genders=normalized_filters.get("genders"),
                classification_filters=normalized_filters.get("classification_filters"),
            )
            df = df.sort_values([self.PANEL_COL, self.TASK_COL]).reset_index(drop=True) if not df.empty else df

        rows_after = int(len(df))
        panelists_after = int(df[self.PANEL_COL].nunique()) if self.PANEL_COL in df.columns and not df.empty else 0

        if filters_active and df.empty:
            return self._build_empty_filtered_json_report(
                study_data=study_data,
                filters_applied=normalized_filters,
                filter_meta={
                    "total_rows_before_filter": rows_before,
                    "total_rows_after_filter": 0,
                    "panelists_before_filter": panelists_before,
                    "panelists_after_filter": 0,
                },
                include_raw_data=include_raw_data,
                analysis_options=analysis_options,
            )
        # 1. Preprocess Data (same as generate_report)
        elements_json = study_data.get("elements", [])
        categories_json = study_data.get("categories", [])
        
        # 2. Element Metadata & Column Mapping (same as generate_report)
        element_meta = []
        for el in elements_json:
            cat_obj = el.get("category", {})
            if not cat_obj and el.get("category_id"):
                found_cat = next((c for c in categories_json if c.get("id") == el.get("category_id")), {})
                cat_obj = found_cat
            
            cat_name = cat_obj.get("name")
            el_name = el.get("name")
            
            if not cat_name or not el_name:
                continue
                
            internal_col_name = f"{cat_name}_{el_name}"
            candidates = [
                internal_col_name,
                f"{cat_name}-{el_name}",
                f"{cat_name}-{el_name}".replace('_', '-').replace(' ', '-'),
                f"{cat_name}_{el_name}".replace(' ', '_')
            ]
            
            actual_col = None
            for cand in candidates:
                if cand in df.columns:
                    actual_col = cand
                    break
            
            if actual_col:
                element_meta.append({
                    "csv_col": actual_col,
                    "category_name": cat_name,
                    "element_name": el_name,
                    "category_order": cat_obj.get("order", 0),
                })

        element_cols_raw = [m["csv_col"] for m in element_meta]
        seen = set()
        element_cols = []
        for col in element_cols_raw:
            if col not in seen:
                element_cols.append(col)
                seen.add(col)
        
        col_to_catname = {m["csv_col"]: m["category_name"] for m in element_meta}
        col_to_eltname = {m["csv_col"]: m["element_name"] for m in element_meta}
        
        cat_order = {}
        for m in element_meta:
            name = m["category_name"]
            order = m["category_order"]
            if name not in cat_order or order < cat_order[name]:
                cat_order[name] = order
        sorted_categories = sorted(cat_order.keys(), key=lambda c: cat_order[c])

        # 3. Classification Columns
        known_cols = {self.PANEL_COL, self.RATING_COL, self.RESPONSE_TIME_COL, 
                      self.GENDER_COL, self.AGE_COL, self.TASK_COL}
        known_cols.update(element_cols)
        
        classification_cols = []
        class_qs = study_data.get("classification_questions", [])
        for q in class_qs:
            q_text = q.get("question_text")
            if q_text and q_text in df.columns:
                classification_cols.append(q_text)
                
        # 4. Run Analysis (same as generate_report)
        coef_table_T = self._run_panel_regressions(df, element_cols, "TOP", analysis_options)
        coef_table_B = self._run_panel_regressions(df, element_cols, "BOTTOM", analysis_options)
        coef_table_R = self._run_panel_regressions(df, element_cols, "RESPONSE", analysis_options)
        
        base_size = df[self.PANEL_COL].nunique()
        
        element_means_T = coef_table_T[element_cols].mean(axis=0).round().astype(int)
        element_means_B = coef_table_B[element_cols].mean(axis=0).round().astype(int)
        element_means_R = coef_table_R[element_cols].mean(axis=0)
        
        gender_groups_T = self._build_gender_groups(coef_table_T, df, element_cols)
        gender_groups_B = self._build_gender_groups(coef_table_B, df, element_cols)
        gender_groups_R = self._build_gender_groups(coef_table_R, df, element_cols)
        
        age_groups_T = self._build_age_groups(coef_table_T, df, element_cols)
        age_groups_B = self._build_age_groups(coef_table_B, df, element_cols)
        age_groups_R = self._build_age_groups(coef_table_R, df, element_cols)
        
        class_groups_T = self._build_class_groups(coef_table_T, df, element_cols, classification_cols)
        class_groups_B = self._build_class_groups(coef_table_B, df, element_cols, classification_cols)
        class_groups_R = self._build_class_groups(coef_table_R, df, element_cols, classification_cols)
        
        intercepts_T = self._run_pooled_regression(df, element_cols, "TOP", analysis_options)
        intercepts_B = self._run_pooled_regression(df, element_cols, "BOTTOM", analysis_options)
        intercepts_R = self._run_pooled_regression(df, element_cols, "RESPONSE", analysis_options)
        
        coef_threshold_T = intercepts_T.get("threshold")
        coef_threshold_B = intercepts_B.get("threshold")
        coef_threshold_R = intercepts_R.get("threshold")
        
        # Clustering for Mindsets
        X_T = coef_table_T[element_cols].to_numpy(dtype=float)
        n_samples = X_T.shape[0]
        
        if n_samples >= 2:
            labels_2_T = self._custom_kmeans_pearson(X_T, k=2, seed=101)
        else:
            labels_2_T = np.zeros(n_samples, dtype=int)
            
        if n_samples >= 3:
            labels_3_T = self._custom_kmeans_pearson(X_T, k=3, seed=202)
        else:
            labels_3_T = np.zeros(n_samples, dtype=int)
        
        # 5. Build JSON structure
        result = {
            "analysis_settings": analysis_options,
        }
        
        # 5a. Front Page
        result["Front Page"] = {
            "Title": study_data.get("title", ""),
            "Background": study_data.get("background", ""),
            "Language": study_data.get("language", ""),
            "Launched At": study_data.get("launched_at", ""),
            "Aspect Ratio": study_data.get("aspect_ratio"),
        }
        
        # 5b. Information Block
        info_block = {
            "Study Title": study_data.get("title", ""),
            "Study Type": study_data.get("study_type", ""),
            "Study Background": study_data.get("background", ""),
            "Aspect Ratio": study_data.get("aspect_ratio"),
            "Categories": []
        }
        
        categories = study_data.get("categories", [])
        elements = study_data.get("elements", [])
        for cat in categories:
            cat_name = cat.get("name", "")
            cat_id = cat.get("id")
            cat_info = {
                "name": cat_name,
                "elements": []
            }
            if cat.get("id") is not None:
                cat_info["id"] = cat.get("id")
                if study_data.get("study_type") == "layer":
                    cat_info["layer_id"] = cat.get("id")
            if cat.get("z_index") is not None:
                cat_info["z_index"] = cat.get("z_index")
            if cat.get("transform") is not None:
                cat_info["transform"] = cat.get("transform")
            if cat.get("order") is not None:
                cat_info["order"] = cat.get("order")
            c_elements = [e for e in elements if e.get("category_id") == cat_id]
            for el in c_elements:
                element_info = {
                    "name": el.get("name", ""),
                    "content": el.get("content", "")
                }
                if el.get("id") is not None:
                    element_info["id"] = el.get("id")
                    if study_data.get("study_type") == "layer":
                        element_info["image_id"] = el.get("id")
                if el.get("category_id") is not None:
                    element_info["category_id"] = el.get("category_id")
                    if study_data.get("study_type") == "layer":
                        element_info["layer_id"] = el.get("category_id")
                for key in (
                    "z_index",
                    "transform",
                    "layer_name",
                    "layer_order",
                    "image_order",
                    "alt_text",
                ):
                    if el.get(key) is not None:
                        element_info[key] = el.get(key)
                cat_info["elements"].append(element_info)
            info_block["Categories"].append(cat_info)
        
        result["Information Block"] = info_block
        
        # 5c. Lightweight dashboard stats used by the analytics page overview.
        # This replaces the need to ship full RawData in the optimized endpoint.
        result["dashboard_summary"] = self._build_dashboard_summary(df, categories)

        if include_raw_data:
            # RawData is useful for full exports/debugging, but it is very large for
            # high-response studies. The optimized endpoint skips this block.
            raw_data_list = []
            for _, row in df.iterrows():
                raw_row = {}
                for col in df.columns:
                    val = row[col]
                    # Handle case where duplicate columns return a Series
                    if isinstance(val, pd.Series):
                        val = val.iloc[0] if not val.empty else None
                    
                    # Check for NA values safely
                    try:
                        is_na = pd.isna(val) if not isinstance(val, (list, dict)) else False
                    except (ValueError, TypeError):
                        is_na = False
                    
                    if is_na:
                        raw_row[col] = None
                    elif isinstance(val, (np.integer, np.int64)):
                        raw_row[col] = int(val)
                    elif isinstance(val, (np.floating, np.float64)):
                        raw_row[col] = float(val)
                    elif isinstance(val, pd.Timestamp):
                        raw_row[col] = val.isoformat()
                    else:
                        raw_row[col] = val
                raw_data_list.append(raw_row)
            result["RawData"] = raw_data_list
        
        # 5d. Overall Sheets
        result["(T) Overall"] = self._build_overall_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            element_means_T, base_size, coef_threshold_T, round_vals=True
        )
        result["(B) Overall"] = self._build_overall_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            element_means_B, base_size, coef_threshold_B, round_vals=True
        )
        result["(R) Overall"] = self._build_overall_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            element_means_R, base_size, coef_threshold_R, round_vals=False
        )
        
        # 5e. Mindsets Sheets
        result["(T) Mindsets"] = self._build_mindset_json(
            coef_table_T, element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, coef_threshold_T, labels_2_T, labels_3_T, round_vals=True
        )
        result["(B) Mindsets"] = self._build_mindset_json(
            coef_table_B, element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, coef_threshold_B, labels_2_T, labels_3_T, round_vals=True
        )
        result["(R) Mindsets"] = self._build_mindset_json(
            coef_table_R, element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, coef_threshold_R, labels_2_T, labels_3_T, round_vals=False
        )
        
        # 5f. Gender Sheets
        result["(T) Gender"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            gender_groups_T, coef_threshold_T, round_vals=True
        )
        result["(B) Gender"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            gender_groups_B, coef_threshold_B, round_vals=True
        )
        result["(R) Gender"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            gender_groups_R, coef_threshold_R, round_vals=False
        )
        
        # 5g. Age Sheets
        result["(T) Age"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            age_groups_T, coef_threshold_T, round_vals=True, segment_order=self.AGE_BINS
        )
        result["(B) Age"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            age_groups_B, coef_threshold_B, round_vals=True, segment_order=self.AGE_BINS
        )
        result["(R) Age"] = self._build_segment_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            age_groups_R, coef_threshold_R, round_vals=False, segment_order=self.AGE_BINS
        )
        
        # 5h. Classification Sheets
        result["(T) Classification Questions"] = self._build_classification_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            class_groups_T, coef_threshold_T, round_vals=True
        )
        result["(B) Classification Questions"] = self._build_classification_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            class_groups_B, coef_threshold_B, round_vals=True
        )
        result["(R) Classification Questions"] = self._build_classification_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            class_groups_R, coef_threshold_R, round_vals=False
        )
        
        # 5i. Combined Sheets
        result["(T) Combined"] = self._build_combined_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, element_means_T, gender_groups_T, age_groups_T, class_groups_T,
            coef_threshold_T, round_vals=True
        )
        result["(B) Combined"] = self._build_combined_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, element_means_B, gender_groups_B, age_groups_B, class_groups_B,
            coef_threshold_B, round_vals=True
        )
        result["(R) Combined"] = self._build_combined_json(
            element_cols, sorted_categories, col_to_catname, col_to_eltname,
            base_size, element_means_R, gender_groups_R, age_groups_R, class_groups_R,
            coef_threshold_R, round_vals=False
        )
        
        # 5j. Intercepts Sheets
        result["(T) Intercepts"] = self._build_intercepts_json(
            intercepts_T["df"], coef_threshold_T,
            intercept=intercepts_T.get("intercept"), t_intercept=intercepts_T.get("t_intercept"),
            include_intercept=self._include_intercept(analysis_options),
        )
        result["(B) Intercepts"] = self._build_intercepts_json(
            intercepts_B["df"], coef_threshold_B,
            intercept=intercepts_B.get("intercept"), t_intercept=intercepts_B.get("t_intercept"),
            include_intercept=self._include_intercept(analysis_options),
        )
        result["(R) Intercepts"] = self._build_intercepts_json(
            intercepts_R["df"], coef_threshold_R,
            intercept=intercepts_R.get("intercept"), t_intercept=intercepts_R.get("t_intercept"),
            include_intercept=self._include_intercept(analysis_options),
        )

        if filters_active:
            self._apply_filter_segment_masks(result, normalized_filters, study_data)
            result["filters_applied"] = normalized_filters
            result["filter_meta"] = {
                "total_rows_before_filter": rows_before,
                "total_rows_after_filter": rows_after,
                "panelists_before_filter": panelists_before,
                "panelists_after_filter": panelists_after,
            }
        
        return result

    def _normalize_filter_age_groups(self, age_groups: Optional[List[str]]) -> List[str]:
        if not age_groups:
            return []
        alias = {"13-18": "13-17"}
        return [alias.get(g, g) for g in age_groups]

    def _normalize_filters_dict(self, filters: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        if not filters:
            return {}
        out: Dict[str, Any] = {}
        if filters.get("age_groups"):
            out["age_groups"] = self._normalize_filter_age_groups(list(filters["age_groups"]))
        if filters.get("genders"):
            out["genders"] = [str(g) for g in filters["genders"]]
        if filters.get("classification_filters"):
            out["classification_filters"] = {
                str(k): [str(v) for v in vals]
                for k, vals in filters["classification_filters"].items()
                if vals
            }
        return out

    def _filters_are_active(self, filters: Dict[str, Any]) -> bool:
        if not filters:
            return False
        if filters.get("age_groups") or filters.get("genders"):
            return True
        class_f = filters.get("classification_filters") or {}
        return any(vals for vals in class_f.values())

    def _build_information_block(self, study_data: Dict[str, Any]) -> Dict[str, Any]:
        info_block = {
            "Study Title": study_data.get("title", ""),
            "Study Type": study_data.get("study_type", ""),
            "Study Background": study_data.get("background", ""),
            "Aspect Ratio": study_data.get("aspect_ratio"),
            "Categories": [],
        }
        categories = study_data.get("categories", [])
        elements = study_data.get("elements", [])
        for cat in categories:
            cat_name = cat.get("name", "")
            cat_id = cat.get("id")
            cat_info: Dict[str, Any] = {"name": cat_name, "elements": []}
            if cat.get("id") is not None:
                cat_info["id"] = cat.get("id")
                if study_data.get("study_type") == "layer":
                    cat_info["layer_id"] = cat.get("id")
            if cat.get("z_index") is not None:
                cat_info["z_index"] = cat.get("z_index")
            if cat.get("transform") is not None:
                cat_info["transform"] = cat.get("transform")
            if cat.get("order") is not None:
                cat_info["order"] = cat.get("order")
            c_elements = [e for e in elements if e.get("category_id") == cat_id]
            for el in c_elements:
                element_info: Dict[str, Any] = {
                    "name": el.get("name", ""),
                    "content": el.get("content", ""),
                }
                if el.get("id") is not None:
                    element_info["id"] = el.get("id")
                    if study_data.get("study_type") == "layer":
                        element_info["image_id"] = el.get("id")
                if el.get("category_id") is not None:
                    element_info["category_id"] = el.get("category_id")
                    if study_data.get("study_type") == "layer":
                        element_info["layer_id"] = el.get("category_id")
                for key in (
                    "z_index", "transform", "layer_name", "layer_order", "image_order", "alt_text",
                ):
                    if el.get(key) is not None:
                        element_info[key] = el.get(key)
                cat_info["elements"].append(element_info)
            info_block["Categories"].append(cat_info)
        return info_block

    def _build_empty_filtered_json_report(
        self,
        study_data: Dict[str, Any],
        filters_applied: Dict[str, Any],
        filter_meta: Dict[str, Any],
        include_raw_data: bool,
        analysis_options: Dict[str, Any],
    ) -> Dict[str, Any]:
        categories = study_data.get("categories", [])
        info_block = self._build_information_block(study_data)
        empty_segment = {"base_size": 0, "threshold": None, "segments": {}, "categories": []}
        empty_class = {"threshold": None, "questions": [], "categories": []}
        empty_mindset = {
            "base_size": 0,
            "threshold": None,
            "groups": {"Total": {"base_size": 0}},
            "categories": [],
        }
        empty_overall = {"base_size": 0, "threshold": None, "categories": []}
        empty_intercepts = {
            "regression_mode": "with_intercept",
            "threshold": None,
            "intercept": None,
            "t_intercept": None,
            "data": [],
        }
        result: Dict[str, Any] = {
            "analysis_settings": analysis_options,
            "filters_applied": filters_applied,
            "filter_meta": {**filter_meta, "error": "No respondents match the applied filters."},
            "Front Page": {
                "Title": study_data.get("title", ""),
                "Background": study_data.get("background", ""),
                "Language": study_data.get("language", ""),
                "Launched At": study_data.get("launched_at", ""),
                "Aspect Ratio": study_data.get("aspect_ratio"),
            },
            "Information Block": info_block,
            "dashboard_summary": self._build_dashboard_summary(pd.DataFrame(), categories),
            "(T) Overall": empty_overall,
            "(B) Overall": empty_overall,
            "(R) Overall": empty_overall,
            "(T) Mindsets": empty_mindset,
            "(B) Mindsets": empty_mindset,
            "(R) Mindsets": empty_mindset,
            "(T) Gender": empty_segment,
            "(B) Gender": empty_segment,
            "(R) Gender": empty_segment,
            "(T) Age": empty_segment,
            "(B) Age": empty_segment,
            "(R) Age": empty_segment,
            "(T) Classification Questions": empty_class,
            "(B) Classification Questions": empty_class,
            "(R) Classification Questions": empty_class,
            "(T) Combined": {"base_size": 0, "threshold": None, "segments": {}, "categories": []},
            "(B) Combined": {"base_size": 0, "threshold": None, "segments": {}, "categories": []},
            "(R) Combined": {"base_size": 0, "threshold": None, "segments": {}, "categories": []},
            "(T) Intercepts": empty_intercepts,
            "(B) Intercepts": empty_intercepts,
            "(R) Intercepts": empty_intercepts,
        }
        if include_raw_data:
            result["RawData"] = []
        self._apply_filter_segment_masks(result, filters_applied, study_data)
        return result

    def _zero_unselected_segment_section(
        self,
        section: Optional[Dict[str, Any]],
        all_segment_keys: List[str],
        selected_keys: List[str],
        round_vals: bool,
    ) -> None:
        if not section or not selected_keys:
            return
        selected_set = set(selected_keys)
        segments = section.setdefault("segments", {})
        for key in all_segment_keys:
            if key not in selected_set:
                segments[key] = {"base_size": 0}
            elif key not in segments:
                segments[key] = {"base_size": 0}

        zero_val = 0
        for cat in section.get("categories") or []:
            for el in cat.get("elements") or []:
                values = el.setdefault("values", {})
                above = el.setdefault("above_threshold", {})
                for key in all_segment_keys:
                    if key not in selected_set:
                        values[key] = zero_val
                        above[key] = False

    def _zero_classification_element_value(
        self, values: Dict[str, Any], key: str, zero_val: int = 0
    ) -> None:
        if key not in values:
            return
        if isinstance(values.get(key), dict):
            values[key] = {"value": zero_val, "above_threshold": False}
        else:
            values[key] = zero_val

    def _mask_classification_section(
        self,
        section: Optional[Dict[str, Any]],
        classification_filters: Dict[str, List[str]],
        round_vals: bool,
    ) -> None:
        """
        When classification_filters is set, only filtered question(s) and their
        selected answer option(s) keep data. All other questions and options are zeroed.
        """
        if not section or not classification_filters:
            return

        filtered_questions = set(classification_filters.keys())
        zero_val = 0

        for q_data in section.get("questions") or []:
            q_text = q_data.get("question_text")
            if not q_text:
                continue

            segments = q_data.setdefault("segments", {})
            key_prefix = f"{q_text}::"

            if q_text not in filtered_questions:
                for ans in list(segments.keys()):
                    segments[ans] = {"base_size": 0}
                for cat in section.get("categories") or []:
                    for el in cat.get("elements") or []:
                        values = el.get("values") or {}
                        for key in list(values.keys()):
                            if key.startswith(key_prefix):
                                self._zero_classification_element_value(values, key, zero_val)
                continue

            allowed = set(classification_filters[q_text])
            for ans in list(segments.keys()):
                if ans not in allowed:
                    segments[ans] = {"base_size": 0}
            for ans in allowed:
                if ans not in segments:
                    segments[ans] = {"base_size": 0}

            for cat in section.get("categories") or []:
                for el in cat.get("elements") or []:
                    values = el.setdefault("values", {})
                    for ans in list(segments.keys()):
                        key = f"{q_text}::{ans}"
                        if ans not in allowed:
                            self._zero_classification_element_value(values, key, zero_val)

    def _mask_combined_classification_section(
        self,
        section: Optional[Dict[str, Any]],
        classification_filters: Dict[str, List[str]],
    ) -> None:
        """Zero classification segments in Combined sheets except filtered Q/options."""
        if not section or not classification_filters:
            return

        filtered_questions = set(classification_filters.keys())
        class_block = (section.get("segments") or {}).get("Classification") or {}

        for q_text, q_info in class_block.items():
            if not isinstance(q_info, dict):
                continue
            answers = q_info.setdefault("answers", {})
            if q_text not in filtered_questions:
                for ans in list(answers.keys()):
                    answers[ans] = {"base_size": 0}
            else:
                allowed = set(classification_filters[q_text])
                for ans in list(answers.keys()):
                    if ans not in allowed:
                        answers[ans] = {"base_size": 0}
                for ans in allowed:
                    if ans not in answers:
                        answers[ans] = {"base_size": 0}

        for cat in section.get("categories") or []:
            for el in cat.get("elements") or []:
                values = el.get("values") or {}
                above = el.setdefault("above_threshold", {})
                for key in list(values.keys()):
                    if not key.startswith("Classification::"):
                        continue
                    parts = key.split("::", 2)
                    if len(parts) != 3:
                        continue
                    _, q_text, ans = parts
                    should_zero = (
                        q_text not in filtered_questions
                        or ans not in set(classification_filters.get(q_text) or [])
                    )
                    if should_zero:
                        values[key] = 0
                        above[key] = False

    def _apply_filter_segment_masks(
        self,
        result: Dict[str, Any],
        filters: Dict[str, Any],
        study_data: Dict[str, Any],
    ) -> None:
        age_sel = filters.get("age_groups") or []
        gender_sel = filters.get("genders") or []
        class_sel = filters.get("classification_filters") or {}

        if age_sel:
            for prefix in ("(T)", "(B)", "(R)"):
                self._zero_unselected_segment_section(
                    result.get(f"{prefix} Age"),
                    list(self.AGE_BINS),
                    age_sel,
                    round_vals=prefix != "(R)",
                )

        if gender_sel:
            for prefix in ("(T)", "(B)", "(R)"):
                self._zero_unselected_segment_section(
                    result.get(f"{prefix} Gender"),
                    ["Male", "Female"],
                    gender_sel,
                    round_vals=prefix != "(R)",
                )

        if class_sel:
            for prefix in ("(T)", "(B)", "(R)"):
                self._mask_classification_section(
                    result.get(f"{prefix} Classification Questions"),
                    class_sel,
                    round_vals=prefix != "(R)",
                )
                self._mask_combined_classification_section(
                    result.get(f"{prefix} Combined"),
                    class_sel,
                )

        if age_sel and result.get("dashboard_summary"):
            dist = result["dashboard_summary"].get("ageDistribution") or []
            allowed = set(age_sel)
            alias = {"13-17": "13-18"}
            display_allowed = allowed | {alias.get(a, a) for a in allowed}
            result["dashboard_summary"]["ageDistribution"] = [
                d for d in dist if d.get("name") in display_allowed or d.get("name") in allowed
            ]

        if gender_sel and result.get("dashboard_summary"):
            dist = result["dashboard_summary"].get("genderDistribution") or []
            allowed = set(gender_sel)
            result["dashboard_summary"]["genderDistribution"] = [
                d for d in dist if d.get("name") in allowed
            ]

    # --- JSON Builder Helpers ---
    def _build_dashboard_summary(self, df: pd.DataFrame, categories: List[Dict[str, Any]]) -> Dict[str, Any]:
        if df is None or df.empty:
            return {
                "totalResponses": 0,
                "uniquePanelists": 0,
                "totalRespondents": 0,
                "avgResponseTime": 0,
                "avgRating": 0,
                "taskCount": 1,
                "categoryCount": len(categories or []),
                "ratingDistribution": [],
                "ageDistribution": [],
                "genderDistribution": [],
                "responseTimeDistribution": [],
                "responseTimeByTask": [],
            }

        task_series = df[self.TASK_COL] if self.TASK_COL in df.columns else pd.Series(dtype=object)
        all_tasks = task_series.dropna().unique().tolist()
        task_count = len(all_tasks) or 1

        panelist_tasks = {}
        if self.PANEL_COL in df.columns:
            task_source = task_series if self.TASK_COL in df.columns else pd.Series([None] * len(df), index=df.index)
            for panelist, task in zip(df[self.PANEL_COL], task_source):
                if pd.isna(panelist):
                    continue
                panelist_key = str(panelist)
                panelist_tasks.setdefault(panelist_key, set())
                if not pd.isna(task):
                    panelist_tasks[panelist_key].add(task)

        total_respondents = sum(1 for tasks in panelist_tasks.values() if len(tasks) == task_count)

        response_times = self._numeric_values(df, self.RESPONSE_TIME_COL)
        ratings = self._numeric_values(df, self.RATING_COL)

        return {
            "totalResponses": int(len(df)),
            "uniquePanelists": int(len(panelist_tasks)),
            "totalRespondents": int(total_respondents),
            "avgResponseTime": sum(response_times) / len(response_times) if response_times else 0,
            "avgRating": sum(ratings) / len(ratings) if ratings else 0,
            "taskCount": int(task_count),
            "categoryCount": int(len(categories or [])),
            "ratingDistribution": self._build_rating_distribution(ratings),
            "ageDistribution": self._build_age_distribution(df),
            "genderDistribution": self._build_segment_participation(df, self.GENDER_COL),
            "responseTimeDistribution": self._build_response_time_distribution(response_times),
            "responseTimeByTask": self._build_response_time_by_task(df),
        }

    def _numeric_values(self, df: pd.DataFrame, column: str) -> List[float]:
        if column not in df.columns:
            return []
        values = []
        for value in df[column].tolist():
            try:
                number = float(value)
            except (TypeError, ValueError):
                continue
            if not np.isnan(number):
                values.append(number)
        return values

    def _build_response_time_distribution(self, response_times: List[float]) -> List[Dict[str, Any]]:
        buckets = [
            {"name": "Fast (<0.5s)", "max": 0.5, "fill": "#22C55E"},
            {"name": "Medium (0.5-1s)", "max": 1, "fill": "#FCCD5B"},
            {"name": "Slow (1-2s)", "max": 2, "fill": "#F7945A"},
            {"name": "Very Slow (>2s)", "max": float("inf"), "fill": "#C04E35"},
        ]
        counts = [0, 0, 0, 0]
        for time_value in response_times:
            for idx, bucket in enumerate(buckets):
                if float(time_value) < bucket["max"]:
                    counts[idx] += 1
                    break
        return [
            {"name": bucket["name"], "value": int(counts[idx]), "fill": bucket["fill"]}
            for idx, bucket in enumerate(buckets)
            if counts[idx] > 0
        ]

    def _build_response_time_by_task(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        if self.TASK_COL not in df.columns or self.RESPONSE_TIME_COL not in df.columns:
            return []
        by_task = {}
        for task, response_time in zip(df[self.TASK_COL], df[self.RESPONSE_TIME_COL]):
            if pd.isna(task):
                continue
            try:
                task_key = int(task)
                time_value = float(response_time)
            except (TypeError, ValueError):
                continue
            if np.isnan(time_value):
                continue
            by_task.setdefault(task_key, []).append(time_value)
        return [
            {"task": task, "avg": sum(times) / len(times), "count": len(times)}
            for task, times in sorted(by_task.items())
        ]

    def _build_rating_distribution(self, ratings: List[float]) -> List[Dict[str, Any]]:
        counts = {}
        for rating in ratings:
            counts[rating] = counts.get(rating, 0) + 1
        return [
            {"name": f"Rating {self._format_distribution_key(key)}", "value": int(value)}
            for key, value in sorted(counts.items())
        ]

    def _build_age_distribution(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        if self.AGE_COL not in df.columns or self.PANEL_COL not in df.columns:
            return []
        order = ["13-18", "18-24", "25-34", "35-44", "45-54", "55-64", "65+", "Under 13", "Unknown"]
        by_range = {label: set() for label in order}
        for age, panelist in zip(df[self.AGE_COL], df[self.PANEL_COL]):
            if pd.isna(panelist):
                continue
            label = self._age_range_label(age)
            by_range.setdefault(label, set()).add(str(panelist))
        return [
            {"name": label, "value": len(by_range[label])}
            for label in order
            if len(by_range.get(label, set())) > 0
        ]

    def _build_segment_participation(self, df: pd.DataFrame, field: str) -> List[Dict[str, Any]]:
        if field not in df.columns or self.PANEL_COL not in df.columns:
            return []
        by_segment = {}
        for segment_value, panelist in zip(df[field], df[self.PANEL_COL]):
            if pd.isna(segment_value) or pd.isna(panelist):
                continue
            segment = str(segment_value)
            by_segment.setdefault(segment, set()).add(str(panelist))
        return [{"name": name, "value": len(panelists)} for name, panelists in by_segment.items()]

    def _age_range_label(self, age_raw: Any) -> str:
        if pd.isna(age_raw):
            return "Unknown"
        age_str = str(age_raw).strip()
        if age_str in {"13-18", "18-24", "25-34", "35-44", "45-54", "55-64", "65+"}:
            return age_str
        try:
            age = float(age_raw)
        except (TypeError, ValueError):
            return "Unknown"
        if age >= 65:
            return "65+"
        if age >= 55:
            return "55-64"
        if age >= 45:
            return "45-54"
        if age >= 35:
            return "35-44"
        if age >= 25:
            return "25-34"
        if age >= 18:
            return "18-24"
        if age >= 13:
            return "13-18"
        return "Under 13"

    def _format_distribution_key(self, value: Any) -> str:
        try:
            number = float(value)
        except (TypeError, ValueError):
            return str(value)
        return str(int(number)) if number.is_integer() else str(number)

    def _build_overall_json(self, element_cols, sorted_cats, col_to_cat, col_to_elt, means, base, threshold, round_vals):
        result = {
            "base_size": int(base),
            "threshold": float(threshold) if threshold is not None else None,
            "categories": []
        }
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                val = means[col]
                if round_vals:
                    val = int(val)
                else:
                    val = float(val)
                
                element_data = {
                    "code": code,
                    "name": col_to_elt.get(col, col),
                    "value": val,
                    "above_threshold": threshold is not None and val >= threshold
                }
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_mindset_json(self, coef_df, element_cols, sorted_cats, col_to_cat, col_to_elt, base, threshold, l2, l3, round_vals):
        counts_2 = np.bincount(l2, minlength=2).tolist()
        counts_3 = np.bincount(l3, minlength=3).tolist()
        
        means_total = coef_df[element_cols].mean(axis=0)
        means_2 = [
            coef_df.iloc[l2 == i][element_cols].mean(axis=0) if np.any(l2 == i) 
            else pd.Series(0, index=element_cols) 
            for i in range(2)
        ]
        means_3 = [
            coef_df.iloc[l3 == i][element_cols].mean(axis=0) if np.any(l3 == i) 
            else pd.Series(0, index=element_cols) 
            for i in range(3)
        ]
        
        result = {
            "base_size": int(base),
            "threshold": float(threshold) if threshold is not None else None,
            "groups": {
                "Total": {"base_size": int(base)},
                "Mindset_2": {
                    f"Mindset_{i+1}_of_2": {"base_size": int(counts_2[i])} 
                    for i in range(2)
                },
                "Mindset_3": {
                    f"Mindset_{i+1}_of_3": {"base_size": int(counts_3[i])} 
                    for i in range(3)
                }
            },
            "categories": []
        }
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                element_name = col_to_elt.get(col, col)
                
                def get_val(mean_series):
                    val = mean_series[col]
                    if round_vals:
                        return int(val)
                    return float(val)
                
                element_data = {
                    "code": code,
                    "name": element_name,
                    "values": {
                        "Total": get_val(means_total),
                        "Mindset_1_of_2": get_val(means_2[0]),
                        "Mindset_2_of_2": get_val(means_2[1]),
                        "Mindset_1_of_3": get_val(means_3[0]),
                        "Mindset_2_of_3": get_val(means_3[1]),
                        "Mindset_3_of_3": get_val(means_3[2])
                    },
                    "above_threshold": {}
                }
                
                # Check threshold for each value
                for key, val in element_data["values"].items():
                    element_data["above_threshold"][key] = threshold is not None and val >= threshold
                
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result


    def run_filtered_regression_report(
        self,
        study_data: Dict[str, Any],
        df: pd.DataFrame,
        filters: Optional[Dict[str, Any]] = None,
        include_per_panelist: bool = False,
        analysis_options: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Run panel regressions (TOP, BOTTOM, RESPONSE) on a filtered subset of the
        respondents DataFrame. Production-ready: accepts the pre-built df and
        returns a JSON-serializable report.

        Args:
            study_data: Study config with categories, elements, classification_questions.
            df: Respondents DataFrame with columns Panelist, Rating, ResponseTime, Task,
                Gender, Age (or AgeGroup), and element columns (CategoryName_ElementName).
                May include classification question columns (question_text as column name).
            filters: Optional dict with:
                - age_groups: list of age bin strings, e.g. ["18-24", "25-34"]
                - genders: list of normalized genders, e.g. ["Male", "Female"]
                - classification_filters: dict { "question_text": ["answer1", "answer2"] }
                Panelists must match all provided filter criteria.
            include_per_panelist: If True, include per-panelist coefficient tables in the report.

        Returns:
            JSON-serializable dict with meta (counts, filters, element_columns), top, bottom,
            response (coefficient_means per mode), and optionally per_panelist.
        """
        filters = filters or {}
        analysis_options = self._resolve_analysis_options(analysis_options)
        age_groups = filters.get("age_groups") or []
        genders = filters.get("genders") or []
        classification_filters = filters.get("classification_filters") or {}

        # Resolve element columns (same logic as generate_json_report)
        element_cols = self._resolve_element_cols(study_data, df)
        if not element_cols:
            return self._empty_filtered_report(
                "No element columns found in DataFrame for this study.", filters
            )

        # Required columns
        required = [self.PANEL_COL, self.RATING_COL, self.RESPONSE_TIME_COL, self.TASK_COL]
        missing = [c for c in required if c not in df.columns]
        if missing:
            return self._empty_filtered_report(
                f"DataFrame missing required columns: {missing}", filters
            )

        rows_before = len(df)
        panelists_before = int(df[self.PANEL_COL].nunique())

        # Filter df
        df_filtered = self._filter_df_by_filters(
            df, age_groups=age_groups, genders=genders, classification_filters=classification_filters
        )
        df_filtered = df_filtered.sort_values([self.PANEL_COL, self.TASK_COL]).reset_index(drop=True)

        rows_after = len(df_filtered)
        panelists_after = int(df_filtered[self.PANEL_COL].nunique())

        if df_filtered.empty or not element_cols:
            return self._empty_filtered_report(
                "No respondents match the applied filters.",
                filters,
                rows_before=rows_before,
                panelists_before=panelists_before,
                rows_after=0,
                panelists_after=0,
                element_columns=element_cols,
            )

        # Run regressions
        coef_T = self._run_panel_regressions(df_filtered, element_cols, "TOP", analysis_options)
        coef_B = self._run_panel_regressions(df_filtered, element_cols, "BOTTOM", analysis_options)
        coef_R = self._run_panel_regressions(df_filtered, element_cols, "RESPONSE", analysis_options)

        means_T = coef_T[element_cols].mean(axis=0)
        means_B = coef_B[element_cols].mean(axis=0)
        means_R = coef_R[element_cols].mean(axis=0)

        def to_json_serializable(val):
            if pd.isna(val):
                return None
            if isinstance(val, (np.integer, np.int64)):
                return int(val)
            if isinstance(val, (np.floating, np.float64)):
                return float(val)
            return val

        def series_to_dict(s, round_vals=False):
            return {k: to_json_serializable(round(v) if round_vals else v) for k, v in s.items()}

        report = {
            "meta": {
                "filters_applied": {
                    "age_groups": list(age_groups),
                    "genders": list(genders),
                    "classification_filters": dict(classification_filters),
                },
                "analysis_settings": analysis_options,
                "total_rows_before_filter": rows_before,
                "total_rows_after_filter": rows_after,
                "panelists_before_filter": panelists_before,
                "panelists_after_filter": panelists_after,
                "element_columns": list(element_cols),
            },
            "top": {
                "coefficient_means": series_to_dict(means_T, round_vals=True),
            },
            "bottom": {
                "coefficient_means": series_to_dict(means_B, round_vals=True),
            },
            "response": {
                "coefficient_means": series_to_dict(means_R, round_vals=False),
            },
        }

        if include_per_panelist:
            report["per_panelist"] = {
                "top": coef_T.to_dict(orient="records"),
                "bottom": coef_B.to_dict(orient="records"),
                "response": coef_R.to_dict(orient="records"),
            }
            # Ensure per_panelist values are JSON-serializable
            for mode in ("top", "bottom", "response"):
                for row in report["per_panelist"][mode]:
                    for k, v in row.items():
                        row[k] = to_json_serializable(v)

        return report

    def _empty_filtered_report(
        self,
        message: str,
        filters: Dict[str, Any],
        rows_before: Optional[int] = None,
        panelists_before: Optional[int] = None,
        rows_after: Optional[int] = None,
        panelists_after: Optional[int] = None,
        element_columns: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        return {
            "meta": {
                "filters_applied": {
                    "age_groups": list(filters.get("age_groups") or []),
                    "genders": list(filters.get("genders") or []),
                    "classification_filters": dict(filters.get("classification_filters") or {}),
                },
                "total_rows_before_filter": rows_before,
                "total_rows_after_filter": rows_after,
                "panelists_before_filter": panelists_before,
                "panelists_after_filter": panelists_after,
                "element_columns": list(element_columns) if element_columns else [],
                "error": message,
            },
            "top": {"coefficient_means": {}},
            "bottom": {"coefficient_means": {}},
            "response": {"coefficient_means": {}},
        }

    def _resolve_element_cols(self, study_data: Dict[str, Any], df: pd.DataFrame) -> List[str]:
        """Resolve element column names that exist in df (same logic as generate_json_report)."""
        elements_json = study_data.get("elements", [])
        categories_json = study_data.get("categories", [])
        element_meta = []
        for el in elements_json:
            cat_obj = el.get("category", {})
            if not cat_obj and el.get("category_id"):
                found_cat = next(
                    (c for c in categories_json if c.get("id") == el.get("category_id")), {}
                )
                cat_obj = found_cat
            cat_name = cat_obj.get("name")
            el_name = el.get("name")
            if not cat_name or not el_name:
                continue
            candidates = [
                f"{cat_name}_{el_name}",
                f"{cat_name}-{el_name}",
                f"{cat_name}-{el_name}".replace("_", "-").replace(" ", "-"),
                f"{cat_name}_{el_name}".replace(" ", "_"),
            ]
            for cand in candidates:
                if cand in df.columns:
                    element_meta.append({"csv_col": cand})
                    break
        seen = set()
        out = []
        for m in element_meta:
            c = m["csv_col"]
            if c not in seen:
                out.append(c)
                seen.add(c)
        return out

    def get_t_overall_scores(self, df: pd.DataFrame, study_data: Dict[str, Any]) -> Dict[str, int]:
        """
        Return T Overall score per element for use in project-level export.
        Uses same logic as generate_report: resolve element cols, run panel regressions (TOP), mean per element.
        Returns dict mapping element column name -> integer score. Empty dict if no elements or df empty.
        """
        if df is None or df.empty:
            return {}
        element_cols = self._resolve_element_cols(study_data, df)
        if not element_cols:
            return {}
        for col in element_cols:
            if col not in df.columns:
                return {}
        coef_table_T = self._run_panel_regressions(df, element_cols, "TOP")
        element_means_T = coef_table_T[element_cols].mean(axis=0).round().astype(int)
        return dict(zip(element_cols, element_means_T.tolist()))

    def _filter_df_by_filters(
        self,
        df: pd.DataFrame,
        age_groups: Optional[List[str]] = None,
        genders: Optional[List[str]] = None,
        classification_filters: Optional[Dict[str, List[str]]] = None,
    ) -> pd.DataFrame:
        """Filter to panelists matching all of age_groups, genders, and classification_filters."""
        if df.empty:
            return df
        age_groups = age_groups or []
        genders = genders or []
        classification_filters = classification_filters or {}

        agg_dict = {}
        if self.PANEL_COL not in df.columns:
            return df
        if self.AGE_COL in df.columns:
            agg_dict[self.AGE_COL] = "first"
        if self.GENDER_COL in df.columns:
            agg_dict[self.GENDER_COL] = "first"
        if not agg_dict:
            first = df.groupby(self.PANEL_COL).size().to_frame("_n")
        else:
            first = df.groupby(self.PANEL_COL).agg(agg_dict)
        if self.AGE_COL in first.columns:
            first = first.rename(columns={self.AGE_COL: "_age"})
        if self.GENDER_COL in first.columns:
            first = first.rename(columns={self.GENDER_COL: "_gender"})
        if "AgeGroup" in df.columns:
            first["_age_bin"] = df.groupby(self.PANEL_COL)["AgeGroup"].first()
        elif "_age" in first.columns:
            first["_age_bin"] = first["_age"].apply(self._normalize_age_to_bin)
        else:
            first["_age_bin"] = None
        if "_gender" in first.columns:
            first["_gender_norm"] = first["_gender"].apply(
                lambda x: self._normalize_gender(x) if isinstance(x, str) else None
            )
        else:
            first["_gender_norm"] = None

        mask = np.ones(len(first), dtype=bool)
        if age_groups and "_age_bin" in first.columns and first["_age_bin"].notna().any():
            mask &= first["_age_bin"].isin(age_groups)
        if genders and "_gender_norm" in first.columns and first["_gender_norm"].notna().any():
            mask &= first["_gender_norm"].isin(genders)
        if classification_filters:
            for q_text, allowed in classification_filters.items():
                if not allowed or q_text not in df.columns:
                    continue
                ans_per_panel = df.dropna(subset=[q_text]).groupby(self.PANEL_COL)[q_text].first()
                allowed_set = set(allowed)
                ok = first.index.isin(ans_per_panel[ans_per_panel.isin(allowed_set)].index)
                mask &= ok

        keep = first.index[mask].tolist()
        return df[df[self.PANEL_COL].isin(keep)].copy()

    def _build_segment_json(self, element_cols, sorted_cats, col_to_cat, col_to_elt, groups, threshold, round_vals, segment_order=None):
        if not groups:
            return {"base_size": 0, "threshold": float(threshold) if threshold is not None else None, "segments": {}, "categories": []}
        
        result = {
            "threshold": float(threshold) if threshold is not None else None,
            "segments": {},
            "categories": []
        }
        
        keys = segment_order if segment_order else sorted(groups.keys())
        
        for k in keys:
            if k in groups:
                result["segments"][k] = {
                    "base_size": groups[k]["base"]
                }
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                element_name = col_to_elt.get(col, col)
                
                element_data = {
                    "code": code,
                    "name": element_name,
                    "values": {},
                    "above_threshold": {}
                }
                
                for k in keys:
                    if k in groups:
                        val = groups[k]["means"][col]
                        if round_vals:
                            val = int(round(val))
                        else:
                            val = float(val)
                        element_data["values"][k] = val
                        element_data["above_threshold"][k] = threshold is not None and val >= threshold
                
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_classification_json(self, element_cols, sorted_cats, col_to_cat, col_to_elt, groups, threshold, round_vals):
        if not groups:
            return {"threshold": float(threshold) if threshold is not None else None, "questions": [], "categories": []}
        
        result = {
            "threshold": float(threshold) if threshold is not None else None,
            "questions": [],
            "categories": []
        }
        
        for q_col, info in groups.items():
            question_data = {
                "question_text": info["question_text"],
                "segments": {}
            }
            
            for ans in info["answer_labels"]:
                question_data["segments"][ans] = {
                    "base_size": info["segments"][ans]["base"]
                }
            
            result["questions"].append(question_data)
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                element_name = col_to_elt.get(col, col)
                
                element_data = {
                    "code": code,
                    "name": element_name,
                    "values": {}
                }
                
                for q_col, info in groups.items():
                    for ans in info["answer_labels"]:
                        key = f"{q_col}::{ans}"
                        val = info["segments"][ans]["means"][col]
                        if round_vals:
                            val = int(round(val))
                        else:
                            val = float(val)
                        element_data["values"][key] = {
                            "value": val,
                            "above_threshold": threshold is not None and val >= threshold
                        }
                
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_combined_json(self, element_cols, sorted_cats, col_to_cat, col_to_elt, base, means, g_groups, a_groups, c_groups, threshold, round_vals):
        result = {
            "base_size": int(base),
            "threshold": float(threshold) if threshold is not None else None,
            "segments": {
                "Overall": {"base_size": int(base)},
                "Gender": {},
                "Age": {},
                "Classification": {}
            },
            "categories": []
        }
        
        # Gender segments
        for g in ["Male", "Female"]:
            if g in g_groups:
                result["segments"]["Gender"][g] = {"base_size": g_groups[g]["base"]}
        
        # Age segments
        for a in self.AGE_BINS:
            if a in a_groups:
                result["segments"]["Age"][a] = {"base_size": a_groups[a]["base"]}
        
        # Classification segments
        for q_col, info in c_groups.items():
            result["segments"]["Classification"][info["question_text"]] = {
                "base_size": 0,
                "answers": {}
            }
            for ans in info["answer_labels"]:
                result["segments"]["Classification"][info["question_text"]]["answers"][ans] = {
                    "base_size": info["segments"][ans]["base"]
                }
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            cat_data = {
                "code": letter,
                "name": cat_name,
                "elements": []
            }
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                element_name = col_to_elt.get(col, col)
                
                def get_val(v):
                    if round_vals:
                        return int(round(v))
                    return float(v)
                
                element_data = {
                    "code": code,
                    "name": element_name,
                    "values": {
                        "Overall": get_val(means[col])
                    },
                    "above_threshold": {
                        "Overall": threshold is not None and get_val(means[col]) >= threshold
                    }
                }
                
                # Gender values
                for g in ["Male", "Female"]:
                    if g in g_groups:
                        val = get_val(g_groups[g]["means"][col])
                        element_data["values"][f"Gender::{g}"] = val
                        element_data["above_threshold"][f"Gender::{g}"] = threshold is not None and val >= threshold
                
                # Age values
                for a in self.AGE_BINS:
                    if a in a_groups:
                        val = get_val(a_groups[a]["means"][col])
                        element_data["values"][f"Age::{a}"] = val
                        element_data["above_threshold"][f"Age::{a}"] = threshold is not None and val >= threshold
                
                # Classification values
                for q_col, info in c_groups.items():
                    for ans in info["answer_labels"]:
                        val = get_val(info["segments"][ans]["means"][col])
                        key = f"Classification::{q_col}::{ans}"
                        element_data["values"][key] = val
                        element_data["above_threshold"][key] = threshold is not None and val >= threshold
                
                cat_data["elements"].append(element_data)
            
            result["categories"].append(cat_data)
        
        return result

    def _build_intercepts_json(self, df, threshold, intercept=None, t_intercept=None, include_intercept=True):
        result = {
            "regression_mode": "with_intercept" if include_intercept else "without_intercept",
            "threshold": float(threshold) if threshold is not None else None,
            "intercept": float(intercept) if intercept is not None else None,
            "t_intercept": float(t_intercept) if t_intercept is not None else None,
            "data": []
        }

        for _, row in df.iterrows():
            row_data = {
                "element": str(row["element"]),
                "beta_with_intercept": float(row["beta_with_intercept"]),
                "t_with_intercept": float(row["t_with_intercept"]),
                "beta": float(row["beta_with_intercept"]),
                "t": float(row["t_with_intercept"]),
                "t_above_2": float(row["t_with_intercept"]) >= 2.0
            }
            result["data"].append(row_data)

        return result

    # --- Regression Helpers ---
    def _run_panel_regressions(
        self,
        df: pd.DataFrame,
        element_cols: List[str],
        mode: str,
        analysis_options: Optional[Dict[str, Any]] = None,
    ) -> pd.DataFrame:
        analysis_options = self._resolve_analysis_options(analysis_options)
        include_intercept = self._include_intercept(analysis_options)
        rows = []
        for pid, g in df.groupby(self.PANEL_COL):
            X = g[element_cols].to_numpy(dtype=float)
            
            if mode == "RESPONSE":
                rt = g[self.RESPONSE_TIME_COL].clip(upper=7.0).to_numpy()
                Y = rt.astype(float)
            else:
                ratings = g[self.RATING_COL].to_numpy()
                Y = self._build_y_vector(ratings, mode, analysis_options)

            n = X.shape[0]
            if include_intercept:
                X_design = np.column_stack([np.ones(n), X])
            else:
                X_design = X
            beta_full, _, _, _ = np.linalg.lstsq(X_design, Y, rcond=None)
            intercept = float(beta_full[0]) if include_intercept else None
            beta = beta_full[1:] if include_intercept else beta_full

            # R2 (with intercept when enabled)
            Y_hat = X_design @ beta_full
            sse = float(np.sum((Y - Y_hat) ** 2))
            y_mean = float(np.mean(Y))
            sst = float(np.sum((Y - y_mean) ** 2))
            r2 = np.nan if sst == 0 else 1.0 - sse / sst

            row = {"Panelist": pid, f"R2_{mode}": r2}
            if include_intercept:
                row[f"Intercept_{mode}"] = intercept
            for col_name, b in zip(element_cols, beta):
                row[col_name] = b
            rows.append(row)
            
        return pd.DataFrame(rows)

    def _run_pooled_regression(
        self,
        df: pd.DataFrame,
        element_cols: List[str],
        mode: str,
        analysis_options: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        analysis_options = self._resolve_analysis_options(analysis_options)
        include_intercept = self._include_intercept(analysis_options)
        X_all = df[element_cols].to_numpy(dtype=float)
        
        if mode == "RESPONSE":
            rt_all = df[self.RESPONSE_TIME_COL].clip(upper=7.0).to_numpy()
            Y_all = rt_all
        else:
            ratings_all = df[self.RATING_COL].to_numpy()
            jitter_seed = 123 if mode == "TOP" else 456
            Y_all = self._build_y_vector(ratings_all, mode, analysis_options, jitter_seed=jitter_seed)

        n, p = X_all.shape
        if include_intercept:
            X_design = np.column_stack([np.ones(n), X_all])
        else:
            X_design = X_all
        beta_full, _, _, _ = np.linalg.lstsq(X_design, Y_all, rcond=None)

        # Stats
        Y_hat = X_design @ beta_full
        e = Y_all - Y_hat
        sse = float(np.sum(e ** 2))
        dof = n - X_design.shape[1]
        sigma2 = sse / dof if dof > 0 else 0.0

        try:
            if dof <= 0:
                t_vals = np.zeros_like(beta_full)
            else:
                XtX_inv = np.linalg.inv(X_design.T @ X_design)
                se = np.sqrt(np.diag(sigma2 * XtX_inv))
                se_safe = np.where(se > 0, se, np.nan)
                with np.errstate(invalid="ignore", divide="ignore"):
                    t_vals = np.where(se_safe > 0, beta_full / se_safe, 0.0)
                t_vals = np.nan_to_num(t_vals, nan=0.0, posinf=0.0, neginf=0.0)
        except Exception:
            t_vals = np.zeros_like(beta_full)

        if include_intercept:
            intercept = float(beta_full[0])
            t_intercept = float(t_vals[0])
            beta_with = beta_full[1:]
            t_elements = t_vals[1:]
        else:
            intercept = None
            t_intercept = None
            beta_with = beta_full
            t_elements = t_vals

        pooled_df = pd.DataFrame({
            "element": element_cols,
            "beta_with_intercept": beta_with,
            "t_with_intercept": t_elements,
        })

        # Threshold from pooled model coefficients
        mask = t_elements >= 2.0
        threshold = float(np.min(beta_with[mask])) if np.any(mask) else None

        return {
            "df": pooled_df,
            "threshold": threshold,
            "intercept": intercept,
            "t_intercept": t_intercept,
        }

    # --- Grouping Helpers ---
    def _normalize_gender(self, val):
        if not isinstance(val, str): return np.nan
        v = val.strip().lower()
        if v.startswith("m"): return "Male"
        if v.startswith("f"): return "Female"
        return np.nan

    def _build_gender_groups(self, coef_table, df, element_cols):
        # Map panelist to gender
        gender_map = df.dropna(subset=[self.GENDER_COL]).groupby(self.PANEL_COL)[self.GENDER_COL].first()
        
        groups = {}
        if gender_map.empty: return groups
        
        coef_with_gender = coef_table.merge(gender_map.rename("Gender"), left_on="Panelist", right_index=True, how="left")
        coef_with_gender["Gender_norm"] = coef_with_gender["Gender"].apply(self._normalize_gender)
        
        for g_name in ["Male", "Female"]:
            sub = coef_with_gender[coef_with_gender["Gender_norm"] == g_name]
            if not sub.empty:
                groups[g_name] = {
                    "base": int(sub["Panelist"].nunique()),
                    "means": sub[element_cols].mean(axis=0)
                }
        return groups

    def _normalize_age_to_bin(self, val):
        if pd.isna(val): return np.nan
        
        # Helper to check bin
        def check_bin(v_str):
            clean = v_str.replace(" ", "")
            for b in self.AGE_BINS:
                if clean == b.replace(" ", ""): return b
            return None

        if isinstance(val, (int, float)):
            age = int(val)
            if 13 <= age <= 17: return "13-17"
            if 18 <= age <= 24: return "18-24"
            if 25 <= age <= 34: return "25-34"
            if 35 <= age <= 44: return "35-44"
            if 45 <= age <= 54: return "45-54"
            if 55 <= age <= 64: return "55-64"
            if age >= 65: return "65+"
            return np.nan
            
        if isinstance(val, str):
            b = check_bin(val)
            if b: return b
            if val.replace(" ", "") == "13-18":
                return "13-17"
            # Try parsing number
            digits = "".join(ch if ch.isdigit() else " " for ch in val)
            parts = digits.split()
            if parts:
                try:
                    return self._normalize_age_to_bin(int(parts[0]))
                except: pass
        return np.nan

    def _build_age_groups(self, coef_table, df, element_cols):
        age_map = df.dropna(subset=[self.AGE_COL]).groupby(self.PANEL_COL)[self.AGE_COL].first()
        groups = {}
        if age_map.empty: return groups
        
        age_bin_map = age_map.apply(self._normalize_age_to_bin).dropna()
        coef_with_age = coef_table.merge(age_bin_map.rename("AgeBin"), left_on="Panelist", right_index=True, how="left")
        
        for bin_label in self.AGE_BINS:
            sub = coef_with_age[coef_with_age["AgeBin"] == bin_label]
            if not sub.empty:
                groups[bin_label] = {
                    "base": int(sub["Panelist"].nunique()),
                    "means": sub[element_cols].mean(axis=0)
                }
        return groups

    def _build_class_groups(self, coef_table, df, element_cols, class_cols):
        groups = {}
        for col_name in class_cols:
            if col_name not in df.columns: 
                continue
            
            filtered_df = df.dropna(subset=[col_name])
            if filtered_df.empty:
                continue
                
            ans_series = filtered_df.groupby(self.PANEL_COL)[col_name].first()
            
            # Skip if empty
            if ans_series.empty:
                continue
            
            # Ensure it's a Series (not DataFrame) - handle edge cases with duplicate column names
            if isinstance(ans_series, pd.DataFrame):
                ans_series = ans_series.iloc[:, 0]
            
            # Use a unique internal column name to avoid conflicts
            coef_with_ans = coef_table.merge(ans_series.rename("_cls_answer_"), left_on="Panelist", right_index=True, how="left")
            
            segs = {}
            answer_labels = []
            unique_opts = ans_series.dropna().unique()
            
            for opt in unique_opts:
                sub = coef_with_ans[coef_with_ans["_cls_answer_"] == opt]
                if not sub.empty:
                    segs[opt] = {
                        "base": int(sub["Panelist"].nunique()),
                        "means": sub[element_cols].mean(axis=0)
                    }
                    answer_labels.append(opt)
            
            if segs:
                groups[col_name] = {
                    "question_text": col_name,
                    "answer_labels": sorted(answer_labels, key=str),
                    "segments": segs
                }
        return groups

    def _custom_kmeans_pearson(self, data, k, max_iters=100, seed=42):
        # Center
        row_means = data.mean(axis=1, keepdims=True)
        centered = data - row_means
        # Normalize
        norms = np.linalg.norm(centered, axis=1, keepdims=True)
        norms[norms == 0] = 1e-9
        normalized = centered / norms
        
        n_samples = normalized.shape[0]
        rng = np.random.default_rng(seed)
        
        indices = rng.choice(n_samples, k, replace=False)
        centroids = normalized[indices]
        labels = np.zeros(n_samples, dtype=int)
        
        for _ in range(max_iters):
            dots = normalized @ centroids.T
            new_labels = np.argmax(dots, axis=1)
            if np.array_equal(new_labels, labels): break
            labels = new_labels
            
            new_centroids = np.zeros_like(centroids)
            for i in range(k):
                points = normalized[labels == i]
                if len(points) > 0:
                    mean_vec = points.mean(axis=0)
                    norm = np.linalg.norm(mean_vec)
                    if norm > 1e-9:
                        new_centroids[i] = mean_vec / norm
                    else:
                        new_centroids[i] = normalized[rng.choice(n_samples)]
                else:
                    new_centroids[i] = normalized[rng.choice(n_samples)]
            centroids = new_centroids
            
        return labels

    # --- Excel Helpers ---
    def _autofit_all_cols(self, ws, max_width=30):
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            if max_len > 0:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

    def _wrap_col_B(self, ws):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v: max_len = max(max_len, len(str(v)))
        ws.column_dimensions["B"].width = min(max_len + 2, self.MAX_WIDTH)
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=2)
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "general", vertical="top", wrap_text=True)

    def _create_front_page(self, wb, data):
        ws = wb.create_sheet("Front Page", 0)
        title = data.get("title", "")
        background = data.get("background", "")
        language = data.get("language", "")
        launched_at = data.get("launched_at", "")
        
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 50
        
        labels = ["Title:", "Background:", "Language:", "Launched At:"]
        values = [title, background, language, launched_at]
        
        thick = Side(border_style="thick", color="000000")
        thin = Side(border_style="thin", color="000000")
        
        for i, (lab, val) in enumerate(zip(labels, values)):
            r = 2 + i
            c_lab = ws.cell(row=r, column=2, value=lab)
            c_val = ws.cell(row=r, column=3, value=val)
            c_lab.font = self.bold_font
            
            top = thick if i == 0 else thin
            bottom = thick if i == len(labels) - 1 else thin
            c_lab.border = Border(top=top, left=thick, right=thin, bottom=bottom)
            c_val.border = Border(top=top, left=thin, right=thick, bottom=bottom)

    def _create_info_block(self, wb, data):
        ws = wb.create_sheet("Information Block", 1)
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 60
        
        row = 1
        def w(l, v, bold=True):
            nonlocal row
            c1 = ws.cell(row=row, column=1, value=l)
            c2 = ws.cell(row=row, column=2, value=v)
            if bold: c1.font = self.bold_font
            c2.alignment = Alignment(wrap_text=True)
            row += 1
            
        w("Study Title", data.get("title", ""))
        w("Study Type", data.get("study_type", ""))
        w("Study Background", data.get("background", ""))
        
        # Categories & Elements
        w("", "")
        categories = data.get("categories", [])
        elements = data.get("elements", [])
        
        for cat in categories:
            cat_name = cat.get("name", "")
            cat_id = cat.get("id")
            w("Study Category", cat_name)
            
            c_elements = [e for e in elements if e.get("category_id") == cat_id]
            for el in c_elements:
                w(f"{cat_name} element", el.get("name", ""))
                w("Element Content", el.get("content", ""))
            w("", "")

    def _create_raw_data_sheet(self, wb, df):
        ws = wb.create_sheet("RawData")
        # Headers
        for c_idx, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=c_idx, value=col).font = self.bold_font
        
        # Data
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    def _create_overall_sheet(self, wb, name, element_cols, sorted_cats, col_to_cat, col_to_elt, means, base, threshold, fill, round_vals):
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["D1"] = "Total"; ws["D1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        ws["D2"] = base
        
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                code = f"{letter}{j}"
                val = means[col]
                if round_vals: val = int(val)
                
                ws.cell(row=row, column=1, value=code)
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                ws.cell(row=row, column=4, value=val)
                
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            formula = f'AND(D{first_val}<>"",D{first_val}>={threshold})'
            ws.conditional_formatting.add(f"D{first_val}:D{last_val}", FormulaRule(formula=[formula], fill=fill))

    def _create_mindset_sheet(self, wb, name, coef_df, element_cols, sorted_cats, col_to_cat, col_to_elt, base, threshold, fill, l2, l3, round_vals):
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        
        col_idx = 4
        col_map = {}
        
        # Total
        ws.cell(row=1, column=col_idx, value="Total").font = self.bold_font
        ws.cell(row=2, column=col_idx, value=base)
        col_map["Total"] = col_idx
        col_idx += 2
        
        # 2 Clusters
        counts_2 = np.bincount(l2, minlength=2)
        for i in range(2):
            ws.cell(row=1, column=col_idx, value=f"Mindset {i+1} of 2").font = self.bold_font
            ws.cell(row=2, column=col_idx, value=counts_2[i])
            col_map[f"2_{i}"] = col_idx
            col_idx += 1
        col_idx += 1
        
        # 3 Clusters
        counts_3 = np.bincount(l3, minlength=3)
        for i in range(3):
            ws.cell(row=1, column=col_idx, value=f"Mindset {i+1} of 3").font = self.bold_font
            ws.cell(row=2, column=col_idx, value=counts_3[i])
            col_map[f"3_{i}"] = col_idx
            col_idx += 1
            
        # Calc means
        means_total = coef_df[element_cols].mean(axis=0)
        means_2 = [coef_df.iloc[l2 == i][element_cols].mean(axis=0) if np.any(l2 == i) else pd.Series(0, index=element_cols) for i in range(2)]
        means_3 = [coef_df.iloc[l3 == i][element_cols].mean(axis=0) if np.any(l3 == i) else pd.Series(0, index=element_cols) for i in range(3)]
        
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                ws.cell(row=row, column=1, value=f"{letter}{j}")
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                
                def put(k, v):
                    if round_vals: v = int(round(v))
                    ws.cell(row=row, column=col_map[k], value=v)
                
                put("Total", means_total[col])
                for k in range(2): put(f"2_{k}", means_2[k][col])
                for k in range(3): put(f"3_{k}", means_3[k][col])
                
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            for c_idx in col_map.values():
                col_let = get_column_letter(c_idx)
                f = f'AND({col_let}{first_val}<>"",{col_let}{first_val}>={threshold})'
                ws.conditional_formatting.add(f"{col_let}{first_val}:{col_let}{last_val}", FormulaRule(formula=[f], fill=fill))

    def _create_segment_sheet(self, wb, name, element_cols, sorted_cats, col_to_cat, col_to_elt, groups, threshold, fill, round_vals, segment_order=None):
        if not groups: return
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        
        col_idx = 4
        col_map = {}
        keys = segment_order if segment_order else sorted(groups.keys())
        
        for k in keys:
            if k in groups:
                ws.cell(row=1, column=col_idx, value=k).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=groups[k]["base"])
                col_map[k] = col_idx
                col_idx += 1
                
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                ws.cell(row=row, column=1, value=f"{letter}{j}")
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                
                for k in keys:
                    if k in groups:
                        val = groups[k]["means"][col]
                        if round_vals: val = int(round(val))
                        ws.cell(row=row, column=col_map[k], value=val)
                
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            for c_idx in col_map.values():
                col_let = get_column_letter(c_idx)
                f = f'AND({col_let}{first_val}<>"",{col_let}{first_val}>={threshold})'
                ws.conditional_formatting.add(f"{col_let}{first_val}:{col_let}{last_val}", FormulaRule(formula=[f], fill=fill))

    def _create_classification_sheet(self, wb, name, element_cols, sorted_cats, col_to_cat, col_to_elt, groups, threshold, fill, round_vals):
        if not groups: return
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        
        col_idx = 4
        col_map = {}
        
        for q_col, info in groups.items():
            ws.cell(row=1, column=col_idx, value=info["question_text"]).font = self.bold_font
            ws.cell(row=1, column=col_idx).fill = self.header_fill
            col_idx += 1
            
            for ans in info["answer_labels"]:
                ws.cell(row=1, column=col_idx, value=ans).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=info["segments"][ans]["base"])
                col_map[(q_col, ans)] = col_idx
                col_idx += 1
            col_idx += 1 # Spacer
            
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                ws.cell(row=row, column=1, value=f"{letter}{j}")
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                
                for (q_col, ans), c_idx in col_map.items():
                    val = groups[q_col]["segments"][ans]["means"][col]
                    if round_vals: val = int(round(val))
                    ws.cell(row=row, column=c_idx, value=val)
                
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            for c_idx in col_map.values():
                col_let = get_column_letter(c_idx)
                f = f'AND({col_let}{first_val}<>"",{col_let}{first_val}>={threshold})'
                ws.conditional_formatting.add(f"{col_let}{first_val}:{col_let}{last_val}", FormulaRule(formula=[f], fill=fill))

    def _create_combined_sheet(self, wb, name, element_cols, sorted_cats, col_to_cat, col_to_elt, base, means, g_groups, a_groups, c_groups, threshold, fill, round_vals):
        ws = wb.create_sheet(name)
        ws["B1"] = "Group"; ws["B1"].font = self.bold_font
        ws["B2"] = "Base Size"; ws["B2"].font = self.bold_font
        
        col_idx = 4
        col_map = {}
        
        # Overall
        ws.cell(row=1, column=col_idx, value="Overall").font = self.bold_font
        ws.cell(row=2, column=col_idx, value=base)
        col_map["Overall"] = col_idx
        col_idx += 1
        
        # Gender
        for g in ["Male", "Female"]:
            if g in g_groups:
                ws.cell(row=1, column=col_idx, value=g).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=g_groups[g]["base"])
                col_map[f"G_{g}"] = col_idx
                col_idx += 1
                
        # Age
        for a in self.AGE_BINS:
            if a in a_groups:
                ws.cell(row=1, column=col_idx, value=a).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=a_groups[a]["base"])
                col_map[f"A_{a}"] = col_idx
                col_idx += 1
                
        # Classification
        for q_col, info in c_groups.items():
            ws.cell(row=1, column=col_idx, value=info["question_text"]).font = self.bold_font
            ws.cell(row=1, column=col_idx).fill = self.header_fill
            col_idx += 1
            for ans in info["answer_labels"]:
                ws.cell(row=1, column=col_idx, value=ans).font = self.bold_font
                ws.cell(row=2, column=col_idx, value=info["segments"][ans]["base"])
                col_map[f"C_{q_col}_{ans}"] = col_idx
                col_idx += 1
            col_idx += 1
            
        row = 5
        first_val = None
        
        for i, cat_name in enumerate(sorted_cats):
            letter = self.letters[i]
            ws.cell(row=row, column=2, value=f"{letter}. {cat_name}").font = self.bold_font
            ws.cell(row=row, column=2).fill = self.header_fill
            row += 1
            
            cols = [c for c in element_cols if col_to_cat.get(c) == cat_name]
            for j, col in enumerate(cols, 1):
                ws.cell(row=row, column=1, value=f"{letter}{j}")
                ws.cell(row=row, column=2, value=col_to_elt.get(col, col))
                
                def p(k, v):
                    if round_vals: v = int(round(v))
                    ws.cell(row=row, column=col_map[k], value=v)
                
                p("Overall", means[col])
                for g in ["Male", "Female"]:
                    if g in g_groups: p(f"G_{g}", g_groups[g]["means"][col])
                for a in self.AGE_BINS:
                    if a in a_groups: p(f"A_{a}", a_groups[a]["means"][col])
                for q_col, info in c_groups.items():
                    for ans in info["answer_labels"]:
                        p(f"C_{q_col}_{ans}", info["segments"][ans]["means"][col])
                        
                if first_val is None: first_val = row
                row += 1
            row += 1
            
        self._wrap_col_B(ws)
        
        if threshold is not None and first_val is not None:
            last_val = row - 2
            for c_idx in col_map.values():
                col_let = get_column_letter(c_idx)
                f = f'AND({col_let}{first_val}<>"",{col_let}{first_val}>={threshold})'
                ws.conditional_formatting.add(f"{col_let}{first_val}:{col_let}{last_val}", FormulaRule(formula=[f], fill=fill))

    def _create_intercepts_sheet(self, wb, name, df, threshold, fill, intercept=None, t_intercept=None):
        ws = wb.create_sheet(name)

        # Headers
        for i, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=i, value=col).font = self.bold_font

        # Data
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        data_last_row = ws.max_row
        self._autofit_all_cols(ws)

        footer_row = data_last_row + 2
        if intercept is not None:
            ws.cell(row=footer_row, column=1, value="Intercept").font = self.bold_font
            ws.cell(row=footer_row, column=2, value=intercept)
            footer_row += 1
        if t_intercept is not None:
            ws.cell(row=footer_row, column=1, value="t_intercept").font = self.bold_font
            ws.cell(row=footer_row, column=2, value=t_intercept)
            footer_row += 1
        if threshold is not None:
            ws.cell(row=footer_row, column=1, value="Threshold").font = self.bold_font
            ws.cell(row=footer_row, column=2, value=threshold)

        # Highlight rows where t_with_intercept >= 2 (column C)
        rule = CellIsRule(operator="greaterThanOrEqual", formula=["2.0"], fill=fill)
        ws.conditional_formatting.add(f"C2:C{data_last_row}", rule)
